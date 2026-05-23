from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from models import db, Subject, Student, Attendance, CIE, student_subject
from forms import SubjectForm, StudentForm, AttendanceForm
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from io import BytesIO
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'lms-dev-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///lms.db').replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()


@app.route('/')
def index():
    subject_count = Subject.query.count()
    student_count = Student.query.count()
    today = date.today()
    today_attendance = Attendance.query.filter_by(date=today).count()
    return render_template('index.html', subject_count=subject_count,
                           student_count=student_count, today_attendance=today_attendance)


@app.route('/subjects')
def subjects():
    all_subjects = Subject.query.all()
    return render_template('subjects.html', subjects=all_subjects)


@app.route('/subjects/add', methods=['GET', 'POST'])
def add_subject():
    form = SubjectForm()
    if form.validate_on_submit():
        section = form.section.data if form.section.data else None
        batch = form.batch.data.strip() if form.batch.data else None
        subject = Subject(name=form.name.data, code=form.code.data,
                          category=form.category.data, section=section, batch=batch)
        db.session.add(subject)
        db.session.commit()
        flash('Subject added successfully!', 'success')
        return redirect(url_for('subjects'))
    return render_template('subject_form.html', form=form, title='Add Subject')


@app.route('/subjects/edit/<int:id>', methods=['GET', 'POST'])
def edit_subject(id):
    subject = Subject.query.get_or_404(id)
    form = SubjectForm(obj=subject)
    if form.validate_on_submit():
        subject.name = form.name.data
        subject.code = form.code.data
        subject.category = form.category.data
        subject.section = form.section.data if form.section.data else None
        subject.batch = form.batch.data.strip() if form.batch.data else None
        db.session.commit()
        flash('Subject updated!', 'success')
        return redirect(url_for('subjects'))
    return render_template('subject_form.html', form=form, title='Edit Subject')


@app.route('/subjects/enroll/<int:id>', methods=['GET', 'POST'])
def bulk_enroll_subject(id):
    subject = Subject.query.get_or_404(id)
    if request.method == 'POST':
        student_ids = request.form.getlist('students')
        subject.students = Student.query.filter(Student.id.in_(student_ids)).all()
        db.session.commit()
        flash(f'Enrolled {len(student_ids)} student(s) in {subject.name}', 'success')
        return redirect(url_for('subjects'))
    enrolled_ids = [s.id for s in subject.students]
    all_students = Student.query.order_by(Student.roll_number).all()
    return render_template('bulk_enroll.html', subject=subject,
                           students=all_students, enrolled_ids=enrolled_ids)


@app.route('/subjects/delete/<int:id>')
def delete_subject(id):
    subject = Subject.query.get_or_404(id)
    Attendance.query.filter_by(subject_id=id).delete()
    db.session.execute(student_subject.delete().where(student_subject.c.subject_id == id))
    db.session.delete(subject)
    db.session.commit()
    flash('Subject deleted!', 'success')
    return redirect(url_for('subjects'))


@app.route('/students')
def students():
    subject_id = request.args.get('subject_id', type=int)
    all_subjects = Subject.query.all()
    if subject_id:
        subject = Subject.query.get_or_404(subject_id)
        all_students = Student.query.filter(Student.subjects.any(id=subject_id)).all()
    else:
        subject = None
        all_students = Student.query.all()
    return render_template('students.html', students=all_students, subjects=all_subjects, selected_subject=subject)


@app.route('/students/add', methods=['GET', 'POST'])
def add_student():
    form = StudentForm()
    if form.validate_on_submit():
        student = Student(name=form.name.data, email=form.email.data,
                          roll_number=form.roll_number.data)
        db.session.add(student)
        db.session.commit()
        flash('Student added!', 'success')
        return redirect(url_for('students'))
    return render_template('student_form.html', form=form, title='Add Student')


@app.route('/students/bulk-upload', methods=['POST'])
def bulk_upload_students():
    if 'file' not in request.files:
        flash('No file selected', 'danger')
        return redirect(url_for('students'))
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('students'))
    if not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)', 'danger')
        return redirect(url_for('students'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        added = 0
        skipped = 0
        errors = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            usn = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ''
            email = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not usn or not name:
                errors.append(f'Row {i}: USN and Full Name are required')
                continue
            if Student.query.filter_by(roll_number=usn).first():
                skipped += 1
                continue
            student = Student(name=name, email=email, roll_number=usn)
            db.session.add(student)
            added += 1
        db.session.commit()
        msg = f'Added {added} student(s)'
        if skipped:
            msg += f', {skipped} skipped (duplicate USN)'
        if errors:
            msg += '. Errors: ' + '; '.join(errors)
        flash(msg, 'success' if added else 'warning')
    except Exception as e:
        flash(f'Error reading file: {str(e)}', 'danger')
    return redirect(url_for('students'))


@app.route('/students/edit/<int:id>', methods=['GET', 'POST'])
def edit_student(id):
    student = Student.query.get_or_404(id)
    form = StudentForm(obj=student)
    if form.validate_on_submit():
        student.name = form.name.data
        student.email = form.email.data
        student.roll_number = form.roll_number.data
        db.session.commit()
        flash('Student updated!', 'success')
        return redirect(url_for('students'))
    return render_template('student_form.html', form=form, title='Edit Student')


@app.route('/students/delete/<int:id>')
def delete_student(id):
    student = Student.query.get_or_404(id)
    Attendance.query.filter_by(student_id=id).delete()
    db.session.execute(student_subject.delete().where(student_subject.c.student_id == id))
    db.session.delete(student)
    db.session.commit()
    flash('Student deleted!', 'success')
    return redirect(url_for('students'))


@app.route('/students/enroll/<int:id>', methods=['GET', 'POST'])
def enroll_student(id):
    student = Student.query.get_or_404(id)
    subjects = Subject.query.all()
    if request.method == 'POST':
        subject_ids = request.form.getlist('subjects')
        student.subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
        db.session.commit()
        flash('Subjects enrolled!', 'success')
        return redirect(url_for('students'))
    enrolled_ids = [s.id for s in student.subjects]
    return render_template('enroll.html', student=student, subjects=subjects,
                           enrolled_ids=enrolled_ids)


@app.route('/students/batch-enroll', methods=['POST'])
def batch_enroll_students():
    student_ids = request.form.getlist('student_ids')
    subject_id = request.form.get('subject_id', type=int)
    if not student_ids or not subject_id:
        flash('Select students and a subject', 'warning')
        return redirect(url_for('students'))
    subject = Subject.query.get_or_404(subject_id)
    students = Student.query.filter(Student.id.in_(student_ids)).all()
    for student in students:
        if subject not in student.subjects:
            student.subjects.append(subject)
    db.session.commit()
    flash(f'Enrolled {len(students)} student(s) in {subject.name}', 'success')
    return redirect(url_for('students'))


CATEGORIES = ['IPCC', 'PCC', 'Department Elective', 'Open Elective', 'Lab', 'Project', 'Remedial']

@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    form = AttendanceForm()
    cat = request.args.get('category', '')
    subjects_q = Subject.query.order_by(Subject.category, Subject.name)
    if cat in CATEGORIES:
        subjects_q = subjects_q.filter_by(category=cat)
    form.subject.choices = [(s.id, f'{s.display_name()} ({s.code})') for s in subjects_q.all()]
    students_data = None
    selected_subject = None
    selected_date = None
    selected_category = cat if cat in CATEGORIES else ''

    if form.validate_on_submit():
        selected_subject = Subject.query.get(form.subject.data)
        selected_date = form.date.data
        students_data = []
        enrolled = Student.query.filter(
            Student.subjects.any(id=form.subject.data)
        ).all()
        for student in enrolled:
            record = Attendance.query.filter_by(
                student_id=student.id,
                subject_id=form.subject.data,
                date=selected_date
            ).first()
            students_data.append({
                'student': student,
                'status': record.status if record else 'Present',
                'attendance_id': record.id if record else None
            })
        return render_template('attendance.html', form=form, students_data=students_data,
                               subject=selected_subject, att_date=selected_date,
                               categories=CATEGORIES, selected_category=selected_category)

    return render_template('attendance.html', form=form, students_data=students_data,
                           categories=CATEGORIES, selected_category=selected_category)


@app.route('/attendance/save-all', methods=['POST'])
def attendance_save_all():
    subject_id = request.form.get('subject_id', type=int)
    att_date_str = request.form.get('date')
    att_date = datetime.strptime(att_date_str, '%Y-%m-%d').date()
    student_ids = request.form.getlist('student_id')
    count = 0

    for i, sid in enumerate(student_ids):
        status = request.form.get(f'status_{i}')
        if not status:
            continue
        record = Attendance.query.filter_by(
            student_id=sid, subject_id=subject_id, date=att_date
        ).first()
        if record:
            record.status = status
        else:
            db.session.add(Attendance(
                student_id=sid, subject_id=subject_id,
                date=att_date, status=status
            ))
        count += 1
    db.session.commit()
    flash(f'Attendance saved for {count} student(s)', 'success')
    return redirect(url_for('attendance'))


@app.route('/attendance/reset', methods=['POST'])
def attendance_reset():
    subject_id = request.form.get('subject_id', type=int)
    att_date_str = request.form.get('date')
    att_date = datetime.strptime(att_date_str, '%Y-%m-%d').date()
    Attendance.query.filter_by(subject_id=subject_id, date=att_date).delete()
    db.session.commit()
    flash('Attendance reset for this date', 'info')
    return redirect(url_for('attendance'))


@app.route('/cie', methods=['GET', 'POST'])
def cie_entry():
    subjects = Subject.query.order_by(Subject.category, Subject.name).all()
    students_data = None
    selected_subject = None

    if request.method == 'POST' and 'subject_id' in request.form:
        subject_id = request.form.get('subject_id', type=int)
        selected_subject = Subject.query.get_or_404(subject_id)
        exam_type = request.form.get('exam_type', 'IA-1')
        students = Student.query.filter(Student.subjects.any(id=subject_id)).all()
        students_data = []
        for student in students:
            record = CIE.query.filter_by(
                student_id=student.id, subject_id=subject_id,
                exam_type=exam_type
            ).first()
            students_data.append({
                'student': student,
                'cie': record,
                'exam_type': exam_type
            })
        return render_template('cie.html', subjects=subjects, students_data=students_data,
                               subject=selected_subject, exam_type=exam_type)

    return render_template('cie.html', subjects=subjects)


@app.route('/cie/save', methods=['POST'])
def cie_save():
    subject_id = request.form.get('subject_id', type=int)
    exam_type = request.form.get('exam_type')
    student_ids = request.form.getlist('student_id')
    marks_list = request.form.getlist('marks')
    max_marks = request.form.get('max_marks', type=float) or 50.0

    for sid, marks in zip(student_ids, marks_list):
        if marks == '':
            continue
        record = CIE.query.filter_by(
            student_id=sid, subject_id=subject_id, exam_type=exam_type
        ).first()
        if record:
            record.marks = float(marks)
            record.max_marks = max_marks
        else:
            db.session.add(CIE(
                student_id=sid, subject_id=subject_id,
                exam_type=exam_type, marks=float(marks),
                max_marks=max_marks
            ))
    db.session.commit()
    flash('CIE marks saved!', 'success')
    return redirect(url_for('cie_entry'))


@app.route('/reports/ipcc/<code>')
def ipcc_combined_report(code):
    subjects = Subject.query.filter_by(code=code).order_by(Subject.section).all()
    if not subjects:
        flash('No IPCC subjects found with this code', 'warning')
        return redirect(url_for('reports'))
    subject = subjects[0]
    all_students = Student.query.filter(
        Student.subjects.any(Subject.code == code)
    ).distinct().order_by(Student.roll_number).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{subject.name} Combined IPCC'

    ws.append(['Roll No', 'Name', 'Email'])
    for s in subjects:
        ws.append([f'--- {s.display_name()} ---'])

        dates = db.session.query(Attendance.date).filter_by(
            subject_id=s.id
        ).distinct().order_by(Attendance.date).all()
        dates = [d[0] for d in dates]

        headers = ['', '', '']
        for d in dates:
            headers.append(d.strftime('%m/%d'))
        headers.append('Present')
        headers.append('Days')
        headers.append('%')
        ws.append(headers)

        for student in all_students:
            enrolled = Student.query.filter(
                Student.id == student.id,
                Student.subjects.any(id=s.id)
            ).count()
            if not enrolled:
                continue
            row = [student.roll_number, student.name, student.email or '']
            present = 0
            for d in dates:
                rec = Attendance.query.filter_by(
                    student_id=student.id, subject_id=s.id, date=d
                ).first()
                status = rec.status if rec else 'A'
                row.append(status[:1])
                if rec and rec.status == 'Present':
                    present += 1
            total = len(dates)
            row.append(present)
            row.append(total)
            row.append(round((present / total * 100) if total > 0 else 0, 1))
            ws.append(row)

        ws.append([])

    cie_types = db.session.query(CIE.exam_type).filter(
        CIE.subject_id.in_([s.id for s in subjects])
    ).distinct().all()
    cie_types = [c[0] for c in cie_types]

    if cie_types:
        ws.append(['=== CIE Marks ==='])
        cie_header = ['Roll No', 'Name', 'Email']
        for s in subjects:
            for ct in cie_types:
                cie_header.append(f'{s.section}-{ct}')
            cie_header.append(f'{s.section}-Total')
        ws.append(cie_header)

        for student in all_students:
            row = [student.roll_number, student.name, student.email or '']
            for s in subjects:
                sec_total = 0
                for ct in cie_types:
                    rec = CIE.query.filter_by(
                        student_id=student.id, subject_id=s.id, exam_type=ct
                    ).first()
                    marks = rec.marks if rec else ''
                    row.append(marks)
                    if marks != '':
                        sec_total += marks
                row.append(sec_total)
            ws.append(row)

    for i, width in enumerate([15, 25, 30] + [10]*50, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else chr(64 + (i-1)//26) + chr(65 + (i-1)%26)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe = f'{subject.name}_IPCC_Combined'.replace(' ', '_')
    return send_file(output, download_name=f'{safe}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reports')
def reports():
    subjects = Subject.query.all()
    return render_template('reports.html', subjects=subjects)


@app.route('/reports/export/<int:subject_id>')
def export_excel(subject_id):
    subject = Subject.query.get_or_404(subject_id)
    students = Student.query.filter(Student.subjects.any(id=subject_id)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{subject.display_name()} Attendance'

    ws.append(['Roll No', 'Name', 'Email'])

    dates = db.session.query(Attendance.date).filter_by(
        subject_id=subject_id
    ).distinct().order_by(Attendance.date).all()
    dates = [d[0] for d in dates]

    headers = ['Roll No', 'Name', 'Email']
    for d in dates:
        headers.append(d.strftime('%Y-%m-%d'))
    headers.append('Total Present')
    headers.append('Total Days')
    headers.append('Percentage')
    ws.append(headers)

    for student in students:
        row = [student.roll_number, student.name, student.email or '']
        present_count = 0
        for d in dates:
            record = Attendance.query.filter_by(
                student_id=student.id, subject_id=subject_id, date=d
            ).first()
            status = record.status if record else 'Absent'
            row.append(status)
            if record and record.status == 'Present':
                present_count += 1
        total_days = len(dates)
        row.append(present_count)
        row.append(total_days)
        row.append(round((present_count / total_days * 100) if total_days > 0 else 0, 2))
        ws.append(row)

    col_widths = [15, 25, 30]
    for d in dates:
        col_widths.append(12)
    col_widths.extend([15, 12, 12])
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = subject.display_name().replace(' ', '_').replace('(', '').replace(')', '')
    return send_file(output, download_name=f'{safe_name}_attendance.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print(f'* LMS running on:')
    print(f'* Local:   http://127.0.0.1:5000')
    print(f'* Network: http://{local_ip}:5000')
    print(f'* Share the Network URL with stakeholders')
    app.run(host='0.0.0.0', port=5000, debug=True)
